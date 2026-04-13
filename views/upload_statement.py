from io import BytesIO
from typing import Tuple

from flask_login import current_user
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.models import Account

ROW_FIELD = set(
    [
        "DIACONT",
        "FECHA",
        "MOVIMIENTO",
        "DESCRIP",
        "DEBE",
        "HABER",
        "SALDO",
        "ORDEN",
        "SERIE",
        "COMPROBANTE",
        "USUARIO",
        "ORIGEN",
        "TRANSA",
        "FECHAMOVI",
        "FECHACONT",
    ]
)


def upload_statement(
    db: Session, update_balance: bool, account_id: str, in_memory_file: BytesIO
) -> Tuple[str, str]:

    summary = '<table><tr><th>Date</th><th>Description</th><th style="text-align: right;">Debit</th><th style="text-align: right;">Credit</th><th style="text-align: right;">Balance</th><th>ID</th></tr>\n'
    # Process the uploaded file

    workbook = load_workbook(filename=in_memory_file, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    headers = [cell for cell in next(ws.iter_rows(values_only=True))]
    if not set(headers).issuperset(ROW_FIELD):
        raise ValueError("Uploaded file is missing required columns.")
    balance = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):

        row_data = dict(zip(headers, row))
        desc = row_data.get("DESCRIP", "No Description")
        if desc == "TOTAL":
            continue

        debe = row_data.get("DEBE", " ")
        if debe == " ":
            debit = 0.0
        else:
            debit = float(debe)
        haber = row_data.get("HABER", " ")
        if haber == " ":
            credit = 0.0
        else:
            credit = float(haber)
        balance = float(row_data.get("SALDO", "0").replace(".", ""))
        fecha = row_data.get("FECHAMOVI")
        if not fecha:
            continue
        summary += f'<tr><td>{fecha.date().isoformat()}</td><td>{desc}</td><td  style="text-align: right;">{debit:,.0f}</td><td style="text-align: right;">{credit:,.0f}</td><td style="text-align: right;">{balance:,.0f}</td>'
        summary += f"<td>{row_data.get("MOVIMIENTO", " ")}</td></tr>\n"

    summary += f"</table>"

    response = f"<p>Account ID: {account_id}<br/>Last balance: {balance:,.0f}</p>"
    if update_balance:
        result = (
            db.query(Account)
            .filter_by(id=account_id, user_id=int(current_user.id))
            .first()
        )
        if result is None:
            response += f"<p>Account {account_id} not found in database.</p>"
        else:
            result.balance = str(balance)
            db.commit()
            response += (
                f"<p>Account {account_id} updated with new balance: {balance:,.0f}</p>"
            )

    return response, summary
