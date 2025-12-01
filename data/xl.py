import glob
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from data.datasource import DataSource
from lib.config import BASE_PATH


def get_table(sheet: str, worksheet: str):
    """
    Function to retrieve data from a Local Excel file.

    :param sheet: Name of the file without extension.
    :param worksheet: The name of the worksheet within the sheet.
    :return: List of lists containing the data from the specified worksheet.
    """
    # Open the Google Sheet by its title
    workbook = load_workbook(filename=BASE_PATH + sheet + ".xlsx", data_only=True)

    # Select the desired worksheet
    ws = workbook[worksheet]
    data = list(ws.iter_rows(values_only=True))
    workbook.close()
    # Get all values from the worksheet
    return data


def get_dict(sheet: str, worksheet: str) -> Dict[str, Any]:
    result = {}

    for row in get_table(sheet, worksheet):
        if len(row) < 2:
            continue
        if row[0] is None:
            continue
        result[row[0].lower()] = row[1]
    return result


def get_sheet_settings(sheet: str) -> Dict[str, Any]:
    """
    Function to retrieve settings from a specific sheet.
    :param sheet: Name of the file without extension.
    :return: Dictionary containing the settings from the "Summary" worksheet.
    """
    settings = get_dict(sheet, "Summary")
    return settings


def discover_assets() -> List[DataSource]:
    """
    Lists all files in a specific Local folder."""

    selected = []
    found = glob.glob(os.path.join(BASE_PATH, "*.xlsx"))
    logging.info(f"Discovered {len(found)} local Excel asset files.")
    for file in found:
        new_ds = DataSource(
            "excel",
            os.path.basename(file).split(".")[0].split(".")[0],
            datetime.fromtimestamp(os.path.getmtime(file)),
        )
        new_ds._get_table = get_table
        new_ds._get_dict = get_dict
        new_ds._get_sheet_settings = get_sheet_settings
        selected.append(new_ds)
    return selected
